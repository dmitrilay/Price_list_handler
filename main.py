# ----------------
# version 0.6 для мвидео
# ________________
import openpyxl
from openpyxl import Workbook
from settings import *
import datetime
import logical as lgp


class ParserFile:
    def __init__(self, path_to_open, path_to_save, db, mg, db_shop):
        self.mg = mg
        self.database = db
        self.path_to_open = path_to_open
        self.path_to_save = path_to_save

        self.price_db = {}
        # self.price_db_exit = {'Наименование': [], 'Цена': []}
        self.price_dict_exit = {}
        self.dict_price_entrance = {}

        self.db_new_price = {}
        self.db_update_price = {}
        self.db_shop = db_shop

    def open_file(self, file=None):
        """
        Открываем файл
        вход артикул, цена, цена2, скидка, скидка2
        """
        if file is None:
            file = self.path_to_open
        print('-- Открываем файл')
        wb = openpyxl.load_workbook(file)  # открываем файл
        sheet = wb.active  # Выбираем активный лист
        rows = sheet.max_row + 1  # cols = self.sheet.max_column
        cols = sheet.max_column + 1

        for row in range(1, rows):
            name = sheet.cell(row=row, column=1).value
            self.dict_price_entrance[name] = []
            for col in range(2, cols):
                self.dict_price_entrance[name].append(sheet.cell(row=row, column=col).value)

    def file_processing(self):
        """
        Подготовака даннх к обработке
        """
        print('-- Обрабатываем файл')
        if self.mg == 'mts.xlsx':
            self.price_dict_exit = lgp.logical_price_processing_mts(self.dict_price_entrance)
        elif self.mg == 'dns.xlsx':
            self.price_dict_exit = lgp.logical_price_processing_dns(self.dict_price_entrance)
        elif self.mg == 'mvideo1.xlsx':
            self.price_dict_exit = lgp.logical_price_processing_mvm(self.dict_price_entrance)

        self.price_dict_entrance = None

    def unloading_from_the_database(self):
        """
        Выгружаем данные из базы даннхы
        """
        print('-- Выгружаю данные из базы')
        with self.database.db:
            for data in self.db_shop.select():
                d = data
                self.price_db[d.name] = [d.price_old, d.price_new, d.date_recording, d.display]

    def find_in_the_database(self):
        """
        Ищем позицию в базе, если находим - проверяем было ли
        изменнеие в цене. Если позиция не найдена то сохраняем
        в словарь для дальнейшей записи.
        """
        db_p = self.price_db
        for key, value in self.price_dict_exit.items():
            price_new = int(value)
            if db_p.get(key):
                price_old = int(db_p[key][0])
                if price_old > price_new:
                    self.db_update_price[key] = price_new
            else:
                self.db_new_price[key] = price_new

    def import_db(self):
        for name, price in self.dict_price_entrance.items():
            self.db_new_price[name] = price[0]

    def writing_database_price(self):
        """
        Сохранение новых данных в базу
        """
        if len(self.db_new_price.items()):
            print('-- Сохраняем новые данные в базу данных')
            dt_now = datetime.datetime.now()
            data = []
            for name, price in self.db_new_price.items():
                data.append(
                    {'name': name, 'price_old': price, 'price_new': 0, 'date_recording': dt_now, 'display': 0, })

            with self.database.db.atomic():
                self.db_shop.insert_many(data).execute()

    def updating_database_price(self):
        """
        Запись обработанных данных в базу. Обнавление цены
        """
        if len(self.db_update_price):
            print('-- Обнавления данных')
            local_dict, local_list = {}, []
            for obj in self.database.PriceParser.select():
                local_dict[obj.name] = obj

            for name, price in self.db_update_price.items():
                local_dict[name].price_new = price
                local_dict[name].date_recording = datetime.datetime.now()
                local_list.append(local_dict[name])

            obj = self.database.PriceParser
            obj.bulk_update(local_list, fields=['price_new', 'date_recording'])

    def writing_file_excel(self):
        """
        Сохраняем данные в формате excel(вместо базы данных)
        """
        print('-- Сохраняем файл')
        wb = Workbook()
        ws = wb.active

        stop_for = len(self.price_dict_exit['Наименование'])
        for i in range(0, stop_for):
            ws.cell(row=i + 1, column=1, value=self.price_dict_exit['Наименование'][i])
            ws.cell(row=i + 1, column=2, value=self.price_dict_exit['Цена'][i])

        wb.save(filename=self.path_to_save)

    def writing_file_excel_db(self, path_to_save='database_excel.xlsx'):
        """
        Сохраняем базу данных в формате excel
        """
        wb = Workbook()
        ws = wb.active
        row = 1
        for key, value in self.price_db.items():
            ws.cell(row=row, column=1, value=key)
            columns = 2
            for i in value:
                ws.cell(row=row, column=columns, value=i)
                columns += 1
            row += 1
        wb.save(filename=path_to_save)

    def delete_db(self):
        self.database.PriceParser.drop_table()
