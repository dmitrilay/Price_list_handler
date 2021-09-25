# ----------------
# version 0.6 для мвидео
# ________________
import openpyxl
from openpyxl import Workbook
import models1 as db1
import models2 as db2
import models3 as db3
import models4 as db4
from settings import *
import datetime

import logical_price_processing as lgp


class ParserFile:
    def __init__(self, path_to_open, path_to_save, db, mg):
        self.mg = mg
        self.db = db
        self.path_to_open = path_to_open
        self.path_to_save = path_to_save
        # self.price_db_new = {'Наименование': [], 'Цена': []}
        self.price_db = {}
        # self.price_db_exit = {'Наименование': [], 'Цена': []}
        # self.price_dict_exit = {'Наименование': [], 'Цена': []}
        self.dict_price_entrance = {}
        # self.price_db_update = {'Наименование': [], 'Цена': []}

    def open_file(self, file=None):
        """
        Открываем файл
        вход артикул, цена, цена2, скидка, скидка2
        """
        if file is None:
            file = self.path_to_open
        print('-- Открываем файл')
        wb = openpyxl.load_workbook(file)  # открываем файл
        sheet = wb.active  # Выбираем активный лист
        rows = sheet.max_row + 1  # cols = self.sheet.max_column
        cols = sheet.max_column + 1

        for row in range(1, rows):
            name = sheet.cell(row=row, column=1).value
            self.dict_price_entrance[name] = []
            for col in range(2, cols):
                self.dict_price_entrance[name].append(sheet.cell(row=row, column=col).value)

    def file_processing(self):
        """
        Подготовака даннх к обработке
        """
        print('-- Обрабатываем файл')

        if self.mg == 'mts.xlsx':
            self.price_dict_exit = lgp.logical_price_processing_mts(self.dict_price_entrance)
        elif self.mg == 'dns.xlsx':
            lgp.logical_price_processing_dns(self.dict_price_entrance)
            # self.price_dict_exit =
        else:
            for i in range(0, len(self.price_dict_entrance['Наименование'])):
                article = self.price_dict_entrance['Наименование'][i]
                price_1 = self.price_dict_entrance['Цена_1'][i]
                price_2 = self.price_dict_entrance['Цена_2'][i]
                discount_1 = self.price_dict_entrance['Скидка_1'][i]
                discount_2 = self.price_dict_entrance['Скидка_2'][i]

                if price_1 is not None:
                    try:
                        if int(price_1) > 0:
                            if int(price_1) > int(price_2) > 0:
                                self.price_dict_exit['Цена'].append(price_2)
                            else:
                                self.price_dict_exit['Цена'].append(price_1)
                            self.price_dict_exit['Наименование'].append(article)

                    except Exception:
                        print('ОШИБКА, текст переменной:', price_1)

        self.price_dict_entrance = None

    def unloading_from_the_database(self):
        """
        Выгружаем данные из базы даннхы
        """
        print('-- Выгружаю данные из базы')
        with self.db.db:
            for data in self.db.PriceParser.select():
                d = data
                self.price_db[d.name] = [d.price_old, d.price_new, d.date_recording, d.display]

    def find_in_the_database(self):
        """
        Ищем позицию в базе, если находим - проверяем было ли
        изменнеие в цене. Если позиция не найдена то сохраняем
        в словарь для дальнейшей записи.
        """
        stop = len(self.price_dict_exit['Наименование'])
        for i in range(0, stop):
            name_1 = self.price_dict_exit['Наименование'][i]
            try:
                index_db = self.price_db['Наименование'].index(name_1)
                price_1 = int(self.price_db['Цена'][index_db])
                price_2 = self.price_dict_exit['Цена'][i]
                if price_2 < price_1:
                    self.checking_data_types(name=name_1,
                                             price=price_2,
                                             record_type='update')
            except Exception:
                name_wr = self.price_dict_exit['Наименование'][i]
                price_wr = self.price_dict_exit['Цена'][i]
                self.checking_data_types(name=name_wr, price=price_wr, record_type='write')

        self.writing_file_db()
        self.updating_the_database_price()

    def checking_data_types(self, name, price, record_type=None):
        """
        Привидение даннх к нужному формамату(int...)
        """
        if record_type == 'write':
            self.price_db_new['Наименование'].append(name)
            self.price_db_new['Цена'].append(int(price))
        if record_type == 'update':
            self.price_db_update['Наименование'].append(name)
            self.price_db_update['Цена'].append(int(price))

    def updating_the_database_price(self, ):
        """
        Запись обработанных данных в базу
        Обнавление цены
        """
        dt_now = datetime.datetime.now()
        # query = PriceParser.update(price_new=0)
        # query.execute()
        print('-- Сохраняем обнавленные данные в базу данных')
        length_for = len(self.price_db_update['Наименование'])
        with self.db.db:
            for i in range(0, length_for):
                name = self.price_db_update['Наименование'][i]
                price = self.price_db_update['Цена'][i]
                product = self.db.PriceParser.get(self.db.PriceParser.name == name)
                product.price_new = price
                product.date_recording = dt_now
                product.save()

    def writing_file_db(self):
        """
        Сохранение новых данных в базу
        """
        data = []
        i2, dt_now = 0, datetime.datetime.now()
        stop_for = len(self.price_db_new['Наименование'])
        if stop_for:
            print('-- Сохраняем новые данные в базу')
        for i in range(0, stop_for):
            name = self.price_db_new['Наименование'][i]
            price = self.price_db_new['Цена'][i]
            data.append({'name': name,
                         'price_old': price,
                         'price_new': 0,
                         'date_recording': dt_now,
                         'display': 0, })

            if i2 == 100 or i == stop_for - 1:
                self.db.PriceParser.insert_many(data).execute()
                i2 = 0
                data = []
            i2 = i2 + 1

    def writing_file_excel(self):
        """
        Сохраняем данные в формате excel(вместо базы данных)
        """
        print('-- Сохраняем файл')
        wb = Workbook()
        ws = wb.active

        stop_for = len(self.price_dict_exit['Наименование'])
        for i in range(0, stop_for):
            ws.cell(row=i + 1, column=1, value=self.price_dict_exit['Наименование'][i])
            ws.cell(row=i + 1, column=2, value=self.price_dict_exit['Цена'][i])

        wb.save(filename=self.path_to_save)

    def writing_file_excel_db(self, path_to_save='database_excel.xlsx'):
        """
        Сохраняем базу данных в формате excel
        """
        wb = Workbook()
        ws = wb.active
        row = 1
        for key, value in self.price_db.items():
            ws.cell(row=row, column=1, value=key)
            columns = 2
            for i in value:
                ws.cell(row=row, column=columns, value=i)
                columns += 1
            row += 1
        wb.save(filename=path_to_save)
